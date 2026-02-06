#!/usr/bin/env python3
"""
Work Summary Tool
Summarizes per-user per-project: which users participated in which projects,
lines contributed and feature points (feat commits) per user per project.
Reuses daily_summary git and config logic.
"""

import os
import sys
import argparse
import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Optional
from collections import defaultdict

import pandas as pd

# Reuse from daily_summary
from daily_summary import (
    load_config,
    find_git_repos,
    get_commits,
    get_commit_type,
)


def _format_commit_list(commit_list: List[Dict]) -> str:
    """Format list of commits as multi-line string: [hash] date - message (one per line)."""
    if not commit_list:
        return ''
    lines = []
    for c in commit_list:
        h = c.get('commit_hash', c.get('full_hash', ''))[:8]
        d = c.get('date', '')
        m = (c.get('message', '') or '').strip()
        lines.append(f"[{h}] {d} - {m}" if d else f"[{h}] {m}")
    return '\n'.join(lines)


def aggregate_by_user_and_repository(
    commits: List[Dict], count_fix_as_feature: bool = False
) -> List[Dict]:
    """
    Group commits by (email, repository) and compute per-user per-repo stats.
    User is identified by email. Returns: email, repository, commit_count, insertions, deletions, net_lines, feature_count, commits (commit detail text).
    """
    key = lambda c: (c.get('email', '').strip() or '(unknown)', c.get('repository', ''))
    by_user_repo = defaultdict(lambda: {
        'total_insertions': 0,
        'total_deletions': 0,
        'feature_count': 0,
        'commit_count': 0,
        'commit_list': [],
    })
    for c in commits:
        user_email, repo = key(c)
        by_user_repo[(user_email, repo)]['commit_count'] += 1
        by_user_repo[(user_email, repo)]['total_insertions'] += c.get('insertions', 0)
        by_user_repo[(user_email, repo)]['total_deletions'] += c.get('deletions', 0)
        by_user_repo[(user_email, repo)]['commit_list'].append(c)
        msg = c.get('message', '') or ''
        t = get_commit_type(msg)
        if t == 'feat' or (count_fix_as_feature and t == 'fix'):
            by_user_repo[(user_email, repo)]['feature_count'] += 1

    rows = []
    for (email, repo), data in sorted(by_user_repo.items()):
        ins, dels = data['total_insertions'], data['total_deletions']
        # Sort commits by date for consistent display
        commit_list = sorted(data['commit_list'], key=lambda x: x.get('date', ''))
        rows.append({
            'email': email,
            'repository': repo,
            'commit_count': data['commit_count'],
            'insertions': ins,
            'deletions': dels,
            'net_lines': ins - dels,
            'feature_count': data['feature_count'],
            'commits': _format_commit_list(commit_list),
        })
    return rows


def aggregate_by_user(detail_rows: List[Dict]) -> List[Dict]:
    """From (email, repository) detail rows, compute per-user (by email) totals."""
    by_user = defaultdict(lambda: {
        'project_count': 0,
        'commit_count': 0,
        'insertions': 0,
        'deletions': 0,
        'feature_count': 0,
    })
    for r in detail_rows:
        email = r.get('email', '')
        by_user[email]['project_count'] += 1
        by_user[email]['commit_count'] += r.get('commit_count', 0)
        by_user[email]['insertions'] += r.get('insertions', 0)
        by_user[email]['deletions'] += r.get('deletions', 0)
        by_user[email]['feature_count'] += r.get('feature_count', 0)

    out = []
    for email in sorted(by_user.keys()):
        d = by_user[email]
        out.append({
            'email': email,
            'project_count': d['project_count'],
            'commit_count': d['commit_count'],
            'insertions': d['insertions'],
            'deletions': d['deletions'],
            'net_lines': d['insertions'] - d['deletions'],
            'feature_count': d['feature_count'],
        })
    return out


def _detail_rows_with_user_separators(
    detail_rows: List[Dict], user_count: int
) -> List[Dict]:
    """When multiple users, insert a blank row between different users (by email) for clearer per-user presentation."""
    if not detail_rows or user_count <= 1:
        return detail_rows
    detail_columns = ['email', 'repository', 'commit_count', 'insertions', 'deletions', 'net_lines', 'feature_count', 'commits']
    blank = {k: '' for k in detail_columns}
    out = []
    prev_email = None
    for r in detail_rows:
        email = r.get('email', '')
        if prev_email is not None and email != prev_email:
            out.append(blank)
        out.append(r)
        prev_email = email
    return out


def generate_work_summary_report(
    detail_rows: List[Dict],
    by_user_rows: List[Dict],
    period: str,
    output_file: str,
    output_format: str = 'excel',
    author_specified: bool = False,
) -> None:
    """Write work summary to Excel (By User, By User & Project, Summary), CSV, or JSON.
    When author_specified is False, sheets are ordered and grouped by user (By User first, detail with separators)."""
    if not detail_rows:
        print("No commits found for the specified period.")
        return

    total_commits = sum(r['commit_count'] for r in detail_rows)
    total_insertions = sum(r['insertions'] for r in detail_rows)
    total_deletions = sum(r['deletions'] for r in detail_rows)
    total_net = total_insertions - total_deletions
    total_features = sum(r['feature_count'] for r in detail_rows)
    user_count = len(by_user_rows)
    project_count = len(set(r['repository'] for r in detail_rows if r.get('repository')))

    summary_data = {
        'period': period,
        'user_count': user_count,
        'project_count': project_count,
        'total_commits': total_commits,
        'total_insertions': total_insertions,
        'total_deletions': total_deletions,
        'total_net_lines': total_net,
        'total_feature_count': total_features,
    }

    if output_format == 'auto':
        ext = Path(output_file).suffix.lower()
        if ext == '.csv':
            output_format = 'csv'
        elif ext == '.json':
            output_format = 'json'
        else:
            output_format = 'excel'

    detail_columns = ['email', 'repository', 'commit_count', 'insertions', 'deletions', 'net_lines', 'feature_count', 'commits']
    # When not filtering by author, present by user: sort by email then repo
    if not author_specified and detail_rows:
        detail_rows_sorted = sorted(detail_rows, key=lambda r: (r.get('email', ''), r.get('repository', '')))
    else:
        detail_rows_sorted = detail_rows
    # Excel only: insert blank row between users when multiple users
    detail_for_excel = _detail_rows_with_user_separators(detail_rows_sorted, user_count)
    df_detail = pd.DataFrame(detail_for_excel)
    df_detail = df_detail[[c for c in detail_columns if c in df_detail.columns]]
    df_detail_export = pd.DataFrame(detail_rows_sorted)
    if not df_detail_export.empty:
        df_detail_export = df_detail_export[[c for c in detail_columns if c in df_detail_export.columns]]
    else:
        df_detail_export = df_detail

    df_user = pd.DataFrame(by_user_rows)
    user_columns = ['email', 'project_count', 'commit_count', 'insertions', 'deletions', 'net_lines', 'feature_count']
    df_user = df_user[[c for c in user_columns if c in df_user.columns]]

    if output_format == 'excel':
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # When not specifying user: put By User first so the form is presented by user
            if not author_specified:
                df_user.to_excel(writer, sheet_name='By User', index=False)
                ws_user = writer.sheets['By User']
                for idx, col in enumerate(df_user.columns, 1):
                    max_len = max(df_user[col].astype(str).map(len).max(), len(col)) + 2
                    ws_user.column_dimensions[get_column_letter(idx)].width = min(max_len, 50)
                ws_user.freeze_panes = 'A2'
                df_detail.to_excel(writer, sheet_name='By User & Project', index=False)
                ws = writer.sheets['By User & Project']
            else:
                df_detail.to_excel(writer, sheet_name='By User & Project', index=False)
                ws = writer.sheets['By User & Project']
                df_user.to_excel(writer, sheet_name='By User', index=False)
                ws_user = writer.sheets['By User']
                for idx, col in enumerate(df_user.columns, 1):
                    max_len = max(df_user[col].astype(str).map(len).max(), len(col)) + 2
                    ws_user.column_dimensions[get_column_letter(idx)].width = min(max_len, 50)
                ws_user.freeze_panes = 'A2'
            for idx, col in enumerate(df_detail.columns, 1):
                if col == 'commits':
                    ws.column_dimensions[get_column_letter(idx)].width = 70
                else:
                    max_len = max(df_detail[col].astype(str).map(len).max(), len(col)) + 2
                    ws.column_dimensions[get_column_letter(idx)].width = min(max_len, 50)
            # Enable wrap for commits column so multi-line commit list displays
            commits_col_idx = list(df_detail.columns).index('commits') + 1 if 'commits' in df_detail.columns else 0
            if commits_col_idx:
                letter = get_column_letter(commits_col_idx)
                for row in range(2, len(df_detail) + 2):
                    cell = ws[f'{letter}{row}']
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.freeze_panes = 'A2'

            summary_df = pd.DataFrame({
                'Metric': [
                    'Period', 'Users', 'Projects', 'Total Commits', 'Total Insertions', 'Total Deletions',
                    'Total Net Lines', 'Total Feature Points (feat)',
                ],
                'Value': [
                    summary_data['period'], summary_data['user_count'], summary_data['project_count'],
                    summary_data['total_commits'], summary_data['total_insertions'], summary_data['total_deletions'],
                    summary_data['total_net_lines'], summary_data['total_feature_count'],
                ],
            })
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            summary_ws = writer.sheets['Summary']
            for idx in range(1, 3):
                summary_ws.column_dimensions[get_column_letter(idx)].width = 25

        print(f"\n{'='*70}")
        print(f"Work summary Excel report: {output_file}")
        print(f"{'='*70}")
        print(f"Users: {user_count} | Projects: {project_count} | Commits: {total_commits} | Insertions: {total_insertions} | Deletions: {total_deletions} | Features: {total_features}")
        print(f"{'='*70}\n")

    elif output_format == 'csv':
        df_detail_export.to_csv(output_file, index=False, encoding='utf-8-sig')
        by_user_path = Path(output_file).with_name(Path(output_file).stem + '.by_user.csv')
        df_user.to_csv(by_user_path, index=False, encoding='utf-8-sig')
        summary_path = Path(output_file).with_suffix('.summary.csv')
        pd.DataFrame([summary_data]).to_csv(summary_path, index=False, encoding='utf-8-sig')
        print(f"\n{'='*70}")
        print(f"Work summary CSV: {output_file}")
        print(f"By user: {by_user_path} | Summary: {summary_path}")
        print(f"{'='*70}")
        print(f"Users: {user_count} | Commits: {total_commits} | Features: {total_features}")
        print(f"{'='*70}\n")

    elif output_format == 'json':
        report = {
            'period': period,
            'summary': summary_data,
            'by_user_and_project': df_detail_export.to_dict('records'),
            'by_user': df_user.to_dict('records'),
        }
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False)
        print(f"\n{'='*70}")
        print(f"Work summary JSON: {output_file}")
        print(f"{'='*70}")
        print(f"Users: {user_count} | Commits: {total_commits} | Features: {total_features}")
        print(f"{'='*70}\n")

    else:
        raise ValueError(f"Unsupported output format: {output_format}")


def main():
    parser = argparse.ArgumentParser(
        description='Work summary by user: which users in which projects, lines and feature points (feat) per user',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python work_summary.py /path/to/work/dir --today
  python work_summary.py /path/to/work/dir --lastweek
  python work_summary.py /path/to/work/dir --start 2024-01-01 --end 2024-01-07 -o work_summary.xlsx
        """,
    )
    parser.add_argument('work_dir', nargs='?', help='Directory containing git repositories')
    parser.add_argument('--today', action='store_true', help='Today\'s commits')
    parser.add_argument('--yesterday', action='store_true', help='Yesterday\'s commits')
    parser.add_argument('--lastweek', action='store_true', help='Last 7 days')
    parser.add_argument('--start', type=str, help='Start date (YYYY-MM-DD)')
    parser.add_argument('--end', type=str, help='End date (YYYY-MM-DD)')
    parser.add_argument('--output', '-o', type=str, help='Output file path')
    parser.add_argument('--author', type=str, help='Filter by author name/email')
    parser.add_argument('--format', '-f', type=str, choices=['excel', 'csv', 'json', 'auto'], default='auto')
    parser.add_argument('--config', '-c', type=str, help='Config file path')
    parser.add_argument('--project-prefix', type=str, default='zgzl', help='Only include projects whose name starts with this prefix (default: zgzl). Use --project-prefix "" to include all.')
    parser.add_argument('--fix-as-feature', action='store_true', help='Count fix commits as feature points')

    args = parser.parse_args()
    config = load_config(args.config)

    if not args.work_dir:
        args.work_dir = config.get('work_dir')
    if not args.work_dir:
        parser.error("work_dir is required (argument or config work_dir)")

    now = datetime.now()
    start_date = None
    end_date = None
    period_name = ""

    if args.today:
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = now
        period_name = "Today"
    elif args.yesterday:
        yesterday = now - timedelta(days=1)
        start_date = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
        period_name = "Yesterday"
    elif args.lastweek:
        end_date = now
        start_date = (now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
        period_name = "Last Week (7 days)"
    elif args.start and args.end:
        try:
            start_date = datetime.strptime(args.start, '%Y-%m-%d')
            end_date = datetime.strptime(args.end, '%Y-%m-%d')
            period_name = f"{args.start} to {args.end}"
        except ValueError:
            print("Error: Invalid date format. Use YYYY-MM-DD")
            sys.exit(1)
    else:
        default_range = config.get('default_time_range', 'today')
        if default_range == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now
            period_name = "Today"
        elif default_range == 'yesterday':
            yesterday = now - timedelta(days=1)
            start_date = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
            period_name = "Yesterday"
        elif default_range == 'lastweek':
            end_date = now
            start_date = (now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
            period_name = "Last Week (7 days)"
        else:
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now
            period_name = "Today"

    output_format = args.format
    if output_format == 'auto':
        if args.output:
            ext = Path(args.output).suffix.lower()
            output_format = 'csv' if ext == '.csv' else 'json' if ext == '.json' else config.get('output_format', 'excel')
        else:
            output_format = config.get('output_format', 'excel')

    if not args.output:
        output_dir = Path(config.get('output_dir', '.'))
        output_dir.mkdir(parents=True, exist_ok=True)
        date_str = start_date.strftime('%Y%m%d')
        if args.yesterday or config.get('default_time_range') == 'yesterday':
            date_str = end_date.strftime('%Y%m%d')
        ext_map = {'excel': '.xlsx', 'csv': '.csv', 'json': '.json'}
        ext = ext_map.get(output_format, '.xlsx')
        args.output = str(output_dir / f"work_summary_{date_str}{ext}")
    else:
        Path(args.output).parent.mkdir(parents=True, exist_ok=True)

    print(f"Scanning for git repositories in: {args.work_dir}")
    repos = find_git_repos(args.work_dir)
    # Default prefix zgzl: only stats projects whose name starts with zgzl; config overrides; --project-prefix "" means no filter
    project_prefix = (args.project_prefix if args.project_prefix is not None else config.get('work_summary_project_prefix', 'zgzl'))
    project_prefix = (project_prefix or '').strip()
    if project_prefix:
        repos = [r for r in repos if os.path.basename(r).startswith(project_prefix)]
    if not repos:
        print("No git repositories found." if not project_prefix else f"No git repositories found with name starting with '{project_prefix}'.")
        sys.exit(1)
    print(f"Found {len(repos)} git repositories" + (f" (name starting with '{project_prefix}')" if project_prefix else ""))

    author_filter = args.author or config.get('author')
    all_commits = []
    for repo in repos:
        repo_name = os.path.basename(repo)
        print(f"Processing: {repo_name}")
        commits = get_commits(repo, start_date, end_date, author_filter, last_commit_hash=None)
        all_commits.extend(commits)

    detail_rows = aggregate_by_user_and_repository(all_commits, count_fix_as_feature=args.fix_as_feature)
    by_user_rows = aggregate_by_user(detail_rows)
    author_specified = bool(author_filter and str(author_filter).strip())
    generate_work_summary_report(detail_rows, by_user_rows, period_name, args.output, output_format, author_specified=author_specified)


if __name__ == '__main__':
    main()
