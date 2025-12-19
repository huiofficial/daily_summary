#!/usr/bin/env python3
"""
Daily Git Commit Summary Tool
Summarizes git commits from multiple repositories into an Excel report.
"""

import os
import sys
import subprocess
import argparse
import json
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Callable
from functools import wraps
import pandas as pd

try:
    import ollama
    OLLAMA_AVAILABLE = True
except ImportError:
    OLLAMA_AVAILABLE = False

try:
    import toml
    TOML_AVAILABLE = True
except ImportError:
    TOML_AVAILABLE = False

# Default configuration
DEFAULT_CONFIG = {
    'default_time_range': 'today',
    'output_dir': '.',
    'output_format': 'excel',
    'author': None,
    'ollama_model': 'qwen3:0.6b',
    'enable_ai_summary': True,
    'retry_attempts': 3,
    'retry_delay': 1,
    'commit_message_template': '{repository} - {author} - {message}',
    'incremental_state_file': '.daily_summary_state.json'
}

CONFIG_FILE_NAME = 'daily_summary_config.toml'


def load_config(config_path: Optional[str] = None) -> Dict:
    """Load configuration from TOML file or return defaults."""
    config = DEFAULT_CONFIG.copy()
    
    if config_path is None:
        # Try to find config in current directory or script directory
        script_dir = Path(__file__).parent
        possible_paths = [
            Path(CONFIG_FILE_NAME),
            script_dir / CONFIG_FILE_NAME,
            Path.home() / CONFIG_FILE_NAME
        ]
        
        for path in possible_paths:
            if path.exists():
                config_path = str(path)
                break
    
    if config_path and Path(config_path).exists():
        try:
            if TOML_AVAILABLE:
                with open(config_path, 'r', encoding='utf-8') as f:
                    file_config = toml.load(f)
                    config.update(file_config)
            else:
                # Fallback to JSON if TOML not available
                with open(config_path, 'r', encoding='utf-8') as f:
                    file_config = json.load(f)
                    config.update(file_config)
        except Exception as e:
            print(f"Warning: Failed to load config file {config_path}: {e}")
    
    return config


def retry_on_failure(max_attempts: int = 3, delay: float = 1.0, backoff: float = 1.0):
    """Decorator to retry a function on failure."""
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            attempts = 0
            current_delay = delay
            
            while attempts < max_attempts:
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    attempts += 1
                    if attempts >= max_attempts:
                        raise e
                    print(f"  Retry {attempts}/{max_attempts} after {current_delay}s...")
                    time.sleep(current_delay)
                    current_delay *= backoff
            return None
        return wrapper
    return decorator


def format_commit_message(commit: Dict, template: str = None) -> str:
    """Format commit message using template."""
    if template is None:
        template = DEFAULT_CONFIG['commit_message_template']
    
    try:
        return template.format(**commit)
    except KeyError as e:
        # If template key not found, use default
        return commit.get('message', '')


def load_incremental_state(state_file: str) -> Dict:
    """Load incremental state from file."""
    if not Path(state_file).exists():
        return {}
    
    try:
        with open(state_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Warning: Failed to load incremental state: {e}")
        return {}


def save_incremental_state(state_file: str, state: Dict):
    """Save incremental state to file."""
    try:
        with open(state_file, 'w', encoding='utf-8') as f:
            json.dump(state, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Warning: Failed to save incremental state: {e}")


def filter_new_commits(commits: List[Dict], last_commit_hash: Optional[str] = None) -> List[Dict]:
    """Filter commits that are new since last run."""
    if last_commit_hash is None:
        return commits
    
    new_commits = []
    found_last_commit = False
    
    for commit in commits:
        if found_last_commit:
            new_commits.append(commit)
        elif commit.get('commit_hash') == last_commit_hash:
            found_last_commit = True
    
    return new_commits


def find_git_repos(base_dir: str) -> List[str]:
    """Find all git repositories in the given directory."""
    repos = []
    base_path = Path(base_dir)
    
    if not base_path.exists():
        print(f"Error: Directory '{base_dir}' does not exist.")
        return repos
    
    # Walk through directory to find .git folders
    for root, dirs, files in os.walk(base_path):
        # Skip hidden directories and common non-repo directories
        dirs[:] = [d for d in dirs if not d.startswith('.') and d not in ['node_modules', 'venv', '__pycache__']]
        
        git_dir = Path(root) / '.git'
        if git_dir.exists() and git_dir.is_dir():
            repos.append(str(root))
    
    return repos


def _execute_git_command(cmd: List[str], cwd: str, timeout: int = 30) -> subprocess.CompletedProcess:
    """Execute git command with retry mechanism."""
    max_attempts = 3
    delay = 1.0
    
    for attempt in range(max_attempts):
        try:
            result = subprocess.run(
                cmd,
                cwd=cwd,
                capture_output=True,
                text=True,
                timeout=timeout
            )
            if result.returncode == 0 or attempt == max_attempts - 1:
                return result
        except subprocess.TimeoutExpired:
            if attempt == max_attempts - 1:
                raise
            print(f"  Timeout, retrying ({attempt + 1}/{max_attempts})...")
            time.sleep(delay)
            delay *= 2
        except Exception as e:
            if attempt == max_attempts - 1:
                raise
            print(f"  Error, retrying ({attempt + 1}/{max_attempts}): {e}")
            time.sleep(delay)
            delay *= 2
    
    return result


def get_commits(repo_path: str, start_date: datetime, end_date: datetime, author: str = None, 
                last_commit_hash: Optional[str] = None) -> List[Dict]:
    """Get commits from a repository within the specified date range."""
    commits = []
    
    try:
        # Format dates for git log
        date_format = "%Y-%m-%d"
        start_str = start_date.strftime(date_format)
        end_str = (end_date + timedelta(days=1)).strftime(date_format)  # Include end date
        
        # Build git log command
        cmd = [
            'git', 'log',
            '--since', start_str,
            '--until', end_str,
            '--pretty=format:%H|%an|%ae|%ad|%s',
            '--date=format:%Y-%m-%d %H:%M:%S'
        ]
        
        if author:
            cmd.extend(['--author', author])
        
        # If incremental mode, only get commits after the last one
        if last_commit_hash:
            # In incremental mode, add commit range (exclude the last commit itself)
            # Date filtering will still apply
            cmd.insert(-1, f'{last_commit_hash}..HEAD')
        
        # Execute git command with retry
        result = _execute_git_command(cmd, repo_path, timeout=30)
        
        if result.returncode != 0:
            if result.stderr:
                print(f"Warning: Error getting commits from {repo_path}: {result.stderr.strip()}")
            return commits
        
        # Parse commit output
        for line in result.stdout.strip().split('\n'):
            if not line:
                continue
            
            parts = line.split('|', 4)
            if len(parts) == 5:
                commit_hash, author_name, author_email, commit_date, message = parts
                
                # Get file changes count with retry
                stats_cmd = ['git', 'show', '--stat', '--format=', commit_hash]
                stats_result = _execute_git_command(stats_cmd, repo_path, timeout=10)
                
                files_changed = 0
                insertions = 0
                deletions = 0
                
                if stats_result.returncode == 0:
                    for stat_line in stats_result.stdout.strip().split('\n'):
                        if 'files changed' in stat_line:
                            # Parse stats: " 3 files changed, 45 insertions(+), 12 deletions(-)"
                            parts = stat_line.split(',')
                            if len(parts) > 0:
                                try:
                                    files_changed = int(parts[0].strip().split()[0])
                                except:
                                    pass
                            if len(parts) > 1:
                                if 'insertion' in parts[1]:
                                    try:
                                        insertions = int(parts[1].strip().split()[0])
                                    except:
                                        pass
                            if len(parts) > 2:
                                if 'deletion' in parts[2]:
                                    try:
                                        deletions = int(parts[2].strip().split()[0])
                                    except:
                                        pass
                
                commit_data = {
                    'repository': os.path.basename(repo_path),
                    'commit_hash': commit_hash[:8],  # Short hash
                    'full_hash': commit_hash,  # Full hash for incremental tracking
                    'author': author_name,
                    'email': author_email,
                    'date': commit_date,
                    'message': message,
                    'files_changed': files_changed,
                    'insertions': insertions,
                    'deletions': deletions,
                }
                commits.append(commit_data)
    
    except subprocess.TimeoutExpired:
        print(f"Warning: Timeout getting commits from {repo_path}")
    except Exception as e:
        print(f"Warning: Error processing {repo_path}: {str(e)}")
    
    return commits


def get_repo_info(repo_path: str) -> Dict:
    """Get repository information like remote URL."""
    info = {
        'remote_url': 'N/A',
        'branch': 'N/A'
    }
    
    try:
        # Get remote URL with retry
        result = _execute_git_command(['git', 'remote', 'get-url', 'origin'], repo_path, timeout=5)
        if result.returncode == 0:
            info['remote_url'] = result.stdout.strip()
        
        # Get current branch with retry
        result = _execute_git_command(['git', 'branch', '--show-current'], repo_path, timeout=5)
        if result.returncode == 0:
            info['branch'] = result.stdout.strip()
    except Exception:
        pass
    
    return info


def summarize_repo_commits(repo_name: str, commit_messages: List[str], model: str = 'qwen3:0.6b') -> str:
    """Use ollama to summarize commit messages for a repository."""
    if not OLLAMA_AVAILABLE:
        return "Ollama not available. Install with: pip install ollama"
    
    if not commit_messages:
        return "No commits to summarize."
    
    try:
        # Combine commit messages (limit to avoid token limit)
        # Take at most the first 50 commit messages to avoid overwhelming the model
        messages_to_summarize = commit_messages[:50]
        messages_text = "\n".join([f"{i+1}. {msg}" for i, msg in enumerate(messages_to_summarize)])
        
        if len(commit_messages) > 50:
            messages_text += f"\n... (还有 {len(commit_messages) - 50} 条提交信息未显示)"
        
        # Create prompt for summarization
        prompt = f"""请客观地提取并整理以下项目 '{repo_name}' 的提交信息内容（用中文）：

提交信息列表：
{messages_text}

要求：
1. 只提取提交信息中的实际工作内容
2. 客观描述做了什么，不使用评价性语言
3. 用2-5句话简洁地整理主要内容
4. 直接陈述事实，不添加"改进"、"优化"、"完善"等评价性词汇"""

        # Call ollama API
        response = ollama.chat(
            model=model,
            messages=[
                {
                    'role': 'user',
                    'content': prompt
                }
            ]
        )
        
        summary = response['message']['content'].strip()
        return summary
    
    except Exception as e:
        error_msg = f"Error generating summary: {str(e)}"
        print(f"    ⚠️  {error_msg}")
        return error_msg


def generate_report(commits: List[Dict], output_file: str, period: str, 
                   repo_summaries: Dict[str, str] = None, output_format: str = 'excel'):
    """Generate report from commits in specified format (excel, csv, json)."""
    if not commits:
        print("No commits found for the specified period.")
        return
    
    # Create DataFrame
    df = pd.DataFrame(commits)
    
    # Reorder columns (exclude full_hash from output)
    columns_order = ['repository', 'date', 'author', 'email', 'commit_hash', 'message', 
                     'files_changed', 'insertions', 'deletions']
    available_columns = [col for col in columns_order if col in df.columns]
    df = df[available_columns]
    
    # Sort by date (newest first)
    df = df.sort_values('date', ascending=False)
    
    # Determine output format from file extension if not specified
    if output_format == 'auto':
        ext = Path(output_file).suffix.lower()
        if ext == '.csv':
            output_format = 'csv'
        elif ext == '.json':
            output_format = 'json'
        else:
            output_format = 'excel'
    
    # Prepare summary data
    summary_data = {
        'period': period,
        'total_commits': len(commits),
        'total_repositories': df['repository'].nunique(),
        'total_files_changed': int(df['files_changed'].sum()),
        'total_insertions': int(df['insertions'].sum()),
        'total_deletions': int(df['deletions'].sum()),
        'unique_authors': df['author'].nunique()
    }
    
    # Generate report based on format
    if output_format == 'excel':
        # Create Excel writer with formatting
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Commits', index=False)
            
            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Commits']
            
            # Auto-adjust column widths
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, Font
            
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                worksheet.column_dimensions[get_column_letter(idx)].width = min(max_length, 50)
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
            
            # Add summary sheet
            summary_df = pd.DataFrame({
                'Metric': ['Period', 'Total Commits', 'Total Repositories', 'Total Files Changed',
                          'Total Insertions', 'Total Deletions', 'Unique Authors'],
                'Value': [summary_data['period'], summary_data['total_commits'], 
                         summary_data['total_repositories'], summary_data['total_files_changed'],
                         summary_data['total_insertions'], summary_data['total_deletions'],
                         summary_data['unique_authors']]
            })
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format summary sheet
            summary_ws = writer.sheets['Summary']
            for idx in range(1, 3):
                summary_ws.column_dimensions[get_column_letter(idx)].width = 25
            
            # Add repository summaries sheet if available
            if repo_summaries:
                summaries_data = []
                for repo_name, summary in repo_summaries.items():
                    summaries_data.append({
                        'Repository': repo_name,
                        'Summary': summary
                    })
                
                summaries_df = pd.DataFrame(summaries_data)
                summaries_df.to_excel(writer, sheet_name='Repository Summaries', index=False)
                
                # Format repository summaries sheet
                summaries_ws = writer.sheets['Repository Summaries']
                summaries_ws.column_dimensions['A'].width = 30
                summaries_ws.column_dimensions['B'].width = 80
                
                # Enable text wrapping for summary column
                for row in range(2, len(summaries_df) + 2):
                    cell = summaries_ws[f'B{row}']
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        print(f"\n{'='*70}")
        print(f"Excel report generated: {output_file}")
        print(f"{'='*70}")
        print(f"Total commits: {len(commits)}")
        print(f"Total repositories: {df['repository'].nunique()}")
        if repo_summaries:
            print(f"Repository summaries generated: {len(repo_summaries)}")
            print(f"\nThe Excel file contains the following sheets:")
            print(f"  1. Commits - Detailed commit information")
            print(f"  2. Summary - Aggregated statistics")
            print(f"  3. Repository Summaries - AI-generated summaries for each repository")
        print(f"{'='*70}\n")
    
    elif output_format == 'csv':
        # Generate CSV report (commits only, summary in separate file)
        df.to_csv(output_file, index=False, encoding='utf-8-sig')
        
        # Save summary and summaries to separate CSV files if needed
        summary_file = Path(output_file).with_suffix('.summary.csv')
        summary_df = pd.DataFrame([summary_data])
        summary_df.to_csv(summary_file, index=False, encoding='utf-8-sig')
        
        if repo_summaries:
            summaries_file = Path(output_file).with_suffix('.summaries.csv')
            summaries_data = [{'Repository': k, 'Summary': v} for k, v in repo_summaries.items()]
            summaries_df = pd.DataFrame(summaries_data)
            summaries_df.to_csv(summaries_file, index=False, encoding='utf-8-sig')
        
        print(f"\n{'='*70}")
        print(f"CSV report generated: {output_file}")
        print(f"Summary file: {summary_file}")
        if repo_summaries:
            print(f"Summaries file: {summaries_file}")
        print(f"{'='*70}")
        print(f"Total commits: {len(commits)}")
        print(f"Total repositories: {df['repository'].nunique()}")
        print(f"{'='*70}\n")
    
    elif output_format == 'json':
        # Generate JSON report
        report_data = {
            'period': period,
            'summary': summary_data,
            'commits': df.to_dict('records'),
        }
        
        if repo_summaries:
            report_data['repository_summaries'] = repo_summaries
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)
        
        print(f"\n{'='*70}")
        print(f"JSON report generated: {output_file}")
        print(f"{'='*70}")
        print(f"Total commits: {len(commits)}")
        print(f"Total repositories: {df['repository'].nunique()}")
        if repo_summaries:
            print(f"Repository summaries generated: {len(repo_summaries)}")
        print(f"{'='*70}\n")
    
    else:
        raise ValueError(f"Unsupported output format: {output_format}")


def main():
    parser = argparse.ArgumentParser(
        description='Summarize daily git commits from multiple repositories into Excel report',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Summarize today's commits
  python daily_summary.py /path/to/work/dir --today

  # Summarize yesterday's commits
  python daily_summary.py /path/to/work/dir --yesterday

  # Summarize last week's commits
  python daily_summary.py /path/to/work/dir --lastweek

  # Custom date range
  python daily_summary.py /path/to/work/dir --start 2024-01-01 --end 2024-01-07
        """
    )
    
    parser.add_argument('work_dir', nargs='?', help='Directory containing git repositories')
    parser.add_argument('--today', action='store_true', help='Summarize today\'s commits')
    parser.add_argument('--yesterday', action='store_true', help='Summarize yesterday\'s commits')
    parser.add_argument('--lastweek', action='store_true', help='Summarize last week\'s commits (7 days)')
    parser.add_argument('--start', type=str, help='Start date (YYYY-MM-DD)')
    parser.add_argument('--end', type=str, help='End date (YYYY-MM-DD)')
    parser.add_argument('--output', '-o', type=str, help='Output file path (default: auto-generated)')
    parser.add_argument('--author', type=str, help='Filter commits by author name/email')
    parser.add_argument('--format', '-f', type=str, choices=['excel', 'csv', 'json', 'auto'], 
                       default='auto', help='Output format (default: auto-detect from extension)')
    parser.add_argument('--config', '-c', type=str, help='Configuration file path')
    parser.add_argument('--incremental', action='store_true', help='Enable incremental mode (only process new commits)')
    parser.add_argument('--template', type=str, help='Commit message template (use {field} placeholders)')
    parser.add_argument('--no-ai-summary', action='store_true', help='Disable AI summary generation')
    
    args = parser.parse_args()
    
    # Load configuration
    config = load_config(args.config)
    
    # Use config defaults if not specified via command line
    if not args.work_dir:
        if 'work_dir' in config:
            args.work_dir = config['work_dir']
        else:
            parser.error("work_dir is required")
    
    # Determine date range
    now = datetime.now()
    start_date = None
    end_date = None
    period_name = ""
    
    if args.today:
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = now
        period_name = "Today"
    elif args.yesterday:
        yesterday = now - timedelta(days=1)
        start_date = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
        period_name = "Yesterday"
    elif args.lastweek:
        end_date = now
        start_date = (now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
        period_name = "Last Week (7 days)"
    elif args.start and args.end:
        try:
            start_date = datetime.strptime(args.start, '%Y-%m-%d')
            end_date = datetime.strptime(args.end, '%Y-%m-%d')
            period_name = f"{args.start} to {args.end}"
        except ValueError:
            print("Error: Invalid date format. Use YYYY-MM-DD")
            sys.exit(1)
    else:
        # Use config default or today
        default_range = config.get('default_time_range', 'today')
        if default_range == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now
            period_name = "Today"
        elif default_range == 'yesterday':
            yesterday = now - timedelta(days=1)
            start_date = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
            period_name = "Yesterday"
        elif default_range == 'lastweek':
            end_date = now
            start_date = (now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
            period_name = "Last Week (7 days)"
        else:
            # Default to today
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now
            period_name = "Today"
    
    # Determine output format
    output_format = args.format
    if output_format == 'auto':
        if args.output:
            ext = Path(args.output).suffix.lower()
            if ext == '.csv':
                output_format = 'csv'
            elif ext == '.json':
                output_format = 'json'
            else:
                output_format = config.get('output_format', 'excel')
        else:
            output_format = config.get('output_format', 'excel')
    
    # Generate output filename if not provided
    if not args.output:
        output_dir = Path(config.get('output_dir', '.'))
        output_dir.mkdir(parents=True, exist_ok=True)
        date_str = start_date.strftime('%Y%m%d')
        if args.yesterday or (config.get('default_time_range') == 'yesterday'):
            date_str = end_date.strftime('%Y%m%d')
        
        ext_map = {'excel': '.xlsx', 'csv': '.csv', 'json': '.json'}
        ext = ext_map.get(output_format, '.xlsx')
        args.output = str(output_dir / f"commit_summary_{date_str}{ext}")
    else:
        # Ensure output directory exists
        output_dir = Path(args.output).parent
        output_dir.mkdir(parents=True, exist_ok=True)
    
    # Load incremental state if enabled
    incremental_state = {}
    state_file = config.get('incremental_state_file', '.daily_summary_state.json')
    if args.incremental:
        incremental_state = load_incremental_state(state_file)
    
    # Find all git repositories
    print(f"Scanning for git repositories in: {args.work_dir}")
    repos = find_git_repos(args.work_dir)
    
    if not repos:
        print("No git repositories found in the specified directory.")
        sys.exit(1)
    
    print(f"Found {len(repos)} git repositories")
    
    # Collect commits from all repositories
    all_commits = []
    repo_commits_map = {}  # Map repo name to list of commits
    author_filter = args.author or config.get('author')
    
    # Get template
    template = args.template or config.get('commit_message_template')
    
    for repo in repos:
        repo_name = os.path.basename(repo)
        print(f"Processing: {repo_name}")
        
        # Get last commit hash for incremental mode
        last_commit_hash = None
        if args.incremental and repo_name in incremental_state:
            last_commit_hash = incremental_state[repo_name].get('last_commit_hash')
            if last_commit_hash:
                print(f"  Incremental mode: processing commits after {last_commit_hash[:8]}")
        
        commits = get_commits(repo, start_date, end_date, author_filter, last_commit_hash)
        
        # Apply template to commit messages if specified
        if template and template != DEFAULT_CONFIG['commit_message_template']:
            for commit in commits:
                commit['formatted_message'] = format_commit_message(commit, template)
                commit['message'] = commit['formatted_message']  # Replace original message
        
        all_commits.extend(commits)
        repo_commits_map[repo_name] = commits
        
        # Update incremental state
        if args.incremental and commits:
            if repo_name not in incremental_state:
                incremental_state[repo_name] = {}
            # Store the most recent commit hash
            incremental_state[repo_name]['last_commit_hash'] = commits[0].get('full_hash', commits[0].get('commit_hash'))
            incremental_state[repo_name]['last_update'] = datetime.now().isoformat()
    
    # Generate summaries for each repository using ollama
    repo_summaries = {}
    enable_ai = config.get('enable_ai_summary', True) and not args.no_ai_summary
    if enable_ai and OLLAMA_AVAILABLE:
        model = config.get('ollama_model', 'qwen3:0.6b')
        print("\n" + "="*70)
        print(f"Generating repository summaries using ollama ({model})...")
        print("="*70)
        for repo_name, commits in repo_commits_map.items():
            if commits:
                print(f"\n[{repo_name}]")
                print(f"  Commits: {len(commits)}")
                commit_messages = [commit.get('message', '') for commit in commits]
                print(f"  Generating summary...")
                summary = summarize_repo_commits(repo_name, commit_messages, model)
                repo_summaries[repo_name] = summary
                print(f"  Summary: {summary}")
        print("\n" + "="*70)
        print("Summary generation completed.")
        print("="*70 + "\n")
    elif enable_ai:
        print("\nWarning: Ollama not available. Skipping repository summaries.")
        print("Install ollama: pip install ollama")
    
    # Save incremental state if enabled
    if args.incremental:
        save_incremental_state(state_file, incremental_state)
    
    # Generate report
    generate_report(all_commits, args.output, period_name, repo_summaries, output_format)


if __name__ == '__main__':
    main()

